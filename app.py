import os
import sys
import importlib.metadata
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_from_directory, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import String
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timezone, timedelta, date
from functools import cmp_to_key
import pandas as pd
from io import BytesIO
from dotenv import load_dotenv

load_dotenv()

def get_brasil_time():
    # Retorna a data/hora atual no fuso horário de Brasília (UTC-3)
    return datetime.now(timezone.utc) - timedelta(hours=3)

def log_change(registration, field_changed, old_value, new_value, change_source='system'):
    """
    Registra uma alteração no log de auditoria
    
    Args:
        registration: Matrícula do funcionário
        field_changed: Nome do campo que foi alterado
        old_value: Valor anterior
        new_value: Novo valor
        change_source: Fonte da alteração ('system' ou 'upload')
    """
    # Converte valores para string para comparação
    def convert_for_comparison(value):
        if isinstance(value, (datetime, date)):
            return value.strftime('%Y-%m-%d')
        return str(value) if value is not None else ''
    
    # Compara os valores convertidos
    if convert_for_comparison(old_value) == convert_for_comparison(new_value):
        return  # Não registra se não houve mudança
        
    # Converte valores para exibição
    if isinstance(old_value, (datetime, date)):
        old_value = old_value.strftime('%d/%m/%Y')
    elif old_value is None:
        old_value = ''
        
    if isinstance(new_value, (datetime, date)):
        new_value = new_value.strftime('%d/%m/%Y')
    elif new_value is None:
        new_value = ''
        
    # Cria o registro de auditoria
    audit_log = AuditLog(
        registration=registration,
        field_changed=field_changed,
        old_value=str(old_value) if old_value is not None else '',
        new_value=str(new_value) if new_value is not None else '',
        changed_by=current_user.username if current_user.is_authenticated else 'system',
        change_source=change_source
    )
    db.session.add(audit_log)

# Correção para Python 3.14
if sys.version_info >= (3, 14):
    import pkgutil
    if not hasattr(pkgutil, 'get_loader'):
        def get_loader(package_name):
            return importlib.util.find_spec(package_name).loader
        pkgutil.get_loader = get_loader

def translate_status(value, field):
    """
    Traduz valores de status do inglês para português
    
    Args:
        value: Valor a ser traduzido
        field: Nome do campo para aplicar a tradução correta
        
    Returns:
        str: Valor traduzido ou o original se não houver tradução
    """
    if not value:
        return value
        
    # Dicionário de tradução para os status
    translations = {
        'status': {
            'active': 'Ativo',
            'inactive': 'Inativo',
            'on_leave': 'Afastado',
            'fired': 'Demitido',
            'demitido': 'Demitido',  # Mantém se já estiver em português
            'desligado': 'Demitido',
            'desligada': 'Demitido',
            'afastado': 'Afastado',
            'afastada': 'Afastado',
            'ativo': 'Ativo',
            'ativa': 'Ativo',
            'inativo': 'Inativo',
            'inativa': 'Inativo'
        },
        'course_status': {
            'not_started': 'Não Iniciado',
            'notstarted': 'Não Iniciado',
            'in_progress': 'Em Andamento',
            'inprogress': 'Em Andamento',
            'andamento': 'Em Andamento',
            'completed': 'Concluído',
            'concluido': 'Concluído',
            'concluída': 'Concluído',
            'delayed': 'Atrasado',
            'atrasado': 'Atrasado',
            'cancelled': 'Cancelado',
            'cancelado': 'Cancelado',
            'cancelada': 'Cancelado'
        },
        'operation_ready': {
            'yes': 'Sim',
            'y': 'Sim',
            's': 'Sim',
            'sim': 'Sim',
            'no': 'Não',
            'n': 'Não',
            'não': 'Não',
            'nao': 'Não'
        },
        'employee_type': {
            'trainee': 'Estagiário',
            'estagiario': 'Estagiário',
            'estagiária': 'Estagiário',
            'temporary': 'Temporário',
            'temporario': 'Temporário',
            'temporária': 'Temporário',
            'intern': 'Estagiário',
            'clt': 'CLT',
            'pj': 'PJ',
            'freelancer': 'Freelancer'
        }
    }
    
    # Converte para string e remove espaços extras
    value_str = str(value).strip()
    value_lower = value_str.lower()
    
    # Verifica se o campo tem tradução definida
    if field in translations:
        # Verifica se o valor exato está no dicionário
        if value_lower in translations[field]:
            return translations[field][value_lower]
            
        # Verifica se o valor (em maiúsculas) já está em português
        pt_values = [v.lower() for v in translations[field].values()]
        if value_lower in pt_values:
            # Retorna com a primeira letra maiúscula para padronizar
            return value_str[0].upper() + value_str[1:].lower()
            
        # Verifica correspondência parcial
        for eng, pt in translations[field].items():
            if eng.lower() in value_lower or pt.lower() in value_lower:
                return pt
    
    # Se não encontrou tradução, retorna o valor original
    return value_str

app = Flask(__name__)
app.jinja_env.filters['translate'] = translate_status

# Função helper para gerar URLs dinamicamente com base na marca
def url_for_brand(endpoint, brand=None, **values):
    """
    Gera uma URL dinamicamente, adicionando o prefixo de marca (_vivo ou _claro) 
    ao endpoint se a marca for fornecida.
    """
    from flask import url_for as flask_url_for, request
    
    # Se brand foi passado explicitamente, usa e remove de values
    brand = values.pop('brand', brand)

    # Se brand não foi fornecido, tenta extrair do contexto (request.path)
    if not brand:
        if request.path.startswith('/vivo/'):
            brand = 'Vivo'
        elif request.path.startswith('/claro/'):
            brand = 'Claro'
        else:
            # Se não for possível inferir pela rota, preferir a marca do usuário
            # autenticado (evita escolher 'vivo' por ordem de fallback)
            try:
                from flask_login import current_user
                if current_user.is_authenticated:
                    brand = getattr(current_user, 'brand', None)
            except Exception:
                brand = None

    # Se temos uma marca, tenta construir o endpoint correspondente primeiro
    if brand:
        brand_suffix = '_vivo' if brand.lower() == 'vivo' else '_claro'
        endpoint_with_brand = endpoint + brand_suffix
        try:
            return flask_url_for(endpoint_with_brand, **values)
        except Exception:
            # se falhar, tenta o endpoint genérico como fallback
            try:
                return flask_url_for(endpoint, **values)
            except Exception:
                raise

    # Se não houve marca no contexto, tentar automaticamente endpoints com sufixos
    # Isto evita BuildError quando templates genéricos são renderizados sem brand
    for suffix in ('_vivo', '_claro'):
        try:
            return flask_url_for(endpoint + suffix, **values)
        except Exception:
            continue

    # Por fim, tenta o endpoint genérico
    return flask_url_for(endpoint, **values)

app.jinja_env.globals['url_for_brand'] = url_for_brand


# Context processor to inject brand and branding color into all templates
@app.context_processor
def inject_brand():
    from flask_login import current_user
    brand = None
    try:
        if current_user.is_authenticated:
            brand = getattr(current_user, 'brand', None)
    except Exception:
        brand = None

    # If not available from user, try to infer from the request path
    try:
        path = request.path
        if not brand:
            if path.startswith('/vivo'):
                brand = 'Vivo'
            elif path.startswith('/claro'):
                brand = 'Claro'
    except Exception:
        pass

    if brand:
        project_title = f"Projeto {brand} Casa Cliente"
        if brand.lower() == 'vivo':
            brand_color = '#6f42c1'  # roxo
        elif brand.lower() == 'claro':
            brand_color = '#dc3545'  # vermelho (bootstrap danger)
        else:
            brand_color = '#0d6efd'
    else:
        project_title = 'Gestão de Colaboradores'
        brand_color = '#0d6efd'

    return dict(brand=brand, project_title=project_title, brand_color=brand_color)

# Filtro para formatar data em português
def formatar_data_ptbr(value, formato='%d/%m/%Y'):
    if not value:
        return ''
    # Mapeamento dos dias da semana e meses
    dias_semana = ['Segunda-feira', 'Terça-feira', 'Quarta-feira', 'Quinta-feira', 'Sexta-feira', 'Sábado', 'Domingo']
    meses = ['janeiro', 'fevereiro', 'março', 'abril', 'maio', 'junho', 'julho', 'agosto', 'setembro', 'outubro', 'novembro', 'dezembro']
    
    # Se o formato contiver %A, substituir pelo dia da semana em português
    if '%A' in formato:
        dia_semana = dias_semana[value.weekday()]
        formato = formato.replace('%A', dia_semana)
    
    # Se o formato contiver %B, substituir pelo mês em português
    if '%B' in formato:
        mes = meses[value.month - 1]
        formato = formato.replace('%B', mes)
    
    # Formatar a data
    return value.strftime(formato)

app.jinja_env.filters['data_ptbr'] = formatar_data_ptbr
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'your-secret-key-here')
# Allow overriding the database via env var DATABASE_URI (for independent instances)
# Usando o caminho relativo para o banco de dados na pasta instance/
import os
instance_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'instance')
os.makedirs(instance_path, exist_ok=True)
db_path = os.path.join(instance_path, 'employees.db')
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', os.getenv('DATABASE_URI', f'sqlite:///{db_path}'))
# Supabase fornece a URL de conexão em `DATABASE_URL`. Mantemos compatibilidade com
# a variável mais antiga `DATABASE_URI` e o fallback para o sqlite local.
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Inicialização das extensões
db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Database Models
class AuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    registration = db.Column(db.String(50), nullable=False, index=True)
    field_changed = db.Column(db.String(50), nullable=False)
    old_value = db.Column(db.String(500))
    new_value = db.Column(db.String(500))
    changed_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    changed_by = db.Column(db.String(80), nullable=False)
    change_source = db.Column(db.String(20), nullable=False)  # 'system' or 'upload'

    def to_dict(self):
        return {
            'id': self.id,
            'registration': self.registration,
            'field_changed': self.field_changed,
            'old_value': self.old_value,
            'new_value': self.new_value,
            'changed_at': self.changed_at,
            'changed_by': self.changed_by,
            'change_source': self.change_source
        }

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), nullable=False)
    password = db.Column(db.String(200), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    access_type = db.Column(db.String(20), nullable=False)  # 'admin' or 'user'
    brand = db.Column(db.String(20), nullable=False, default='Vivo')  # 'Vivo' or 'Claro'
    __table_args__ = (db.UniqueConstraint('username', 'brand', name='uq_username_brand'),)

class Employee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    registration = db.Column(db.String(50), nullable=False)
    brand = db.Column(db.String(20), nullable=False, default='Vivo')  # 'Vivo' or 'Claro'
    full_name = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(100))
    employee_type = db.Column(db.String(50))
    admission_date = db.Column(db.Date)
    cep = db.Column(db.String(20))
    status = db.Column(db.String(50))
    course_status = db.Column(db.String(50))
    team = db.Column(db.String(100))
    course_location = db.Column(db.String(200))
    manager = db.Column(db.String(100))
    corporate_manager = db.Column(db.String(100))
    instructor = db.Column(db.String(100))
    contato = db.Column(db.String(20))
    operation_ready = db.Column(String(10))
    integration_start = db.Column(db.Date)
    integration_end = db.Column(db.Date)
    normative_start = db.Column(db.Date)
    normative_end = db.Column(db.Date)
    technical_course_start = db.Column(db.Date)
    technical_course_end = db.Column(db.Date)
    double_start = db.Column(db.Date)
    double_end = db.Column(db.Date)
    loading_date = db.Column(db.Date)
    field_operation_date = db.Column(db.Date)
    last_updated = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)
    __table_args__ = (db.UniqueConstraint('registration', 'brand', name='uq_registration_brand'),)

    def get_current_phase(self):
        # Corrigido erro de comparação datetime vs date - v2
        today = datetime.now().date()
        
        # Função helper para garantir que as datas sejam do tipo date (não datetime)
        def to_date(dt):
            if dt is None:
                return None
            if isinstance(dt, datetime):
                return dt.date()
            return dt
        
        # Converter TODAS as datas relevantes para date antes de qualquer comparação
        integration_start = to_date(self.integration_start)
        integration_end = to_date(self.integration_end)
        normative_start = to_date(self.normative_start)
        normative_end = to_date(self.normative_end)
        technical_course_start = to_date(self.technical_course_start)
        technical_course_end = to_date(self.technical_course_end)
        double_start = to_date(self.double_start)
        double_end = to_date(self.double_end)
        loading_date = to_date(self.loading_date)
        field_op_date = to_date(self.field_operation_date)
        
        # Log detalhado para depuração
        print(f"[LOG FASE] Matrícula: {self.registration} | Nome: {self.full_name}")
        print(f"[LOG FASE] Datas: Integração=({integration_start}, {integration_end}), Normativo=({normative_start}, {normative_end}), Técnico=({technical_course_start}, {technical_course_end}), Duplado=({double_start}, {double_end})")
        print(f"[LOG FASE] Data Carregamento: {loading_date} | Data Operação Campo: {field_op_date}")
        print(f"[LOG FASE] Status do Curso: {self.course_status} | Operation Ready: {self.operation_ready}")
        print(f"[LOG FASE] Data atual: {today}")

        # 1. Verificar se está em alguma fase ativa
        if integration_start and integration_end and integration_start <= today <= integration_end:
            print(f"[LOG FASE] Fase ativa detectada: Integração")
            print(f"[LOG FASE] Fase retornada: Integração\n")
            return 'Integração'
            
        if normative_start and normative_end and normative_start <= today <= normative_end:
            print(f"[LOG FASE] Fase ativa detectada: Normativo")
            print(f"[LOG FASE] Fase retornada: Normativo\n")
            return 'Normativo'
            
        if technical_course_start and technical_course_end and technical_course_start <= today <= technical_course_end:
            print(f"[LOG FASE] Fase ativa detectada: Curso Técnico")
            print(f"[LOG FASE] Fase retornada: Curso Técnico\n")
            return 'Curso Técnico'
            
        if double_start and double_end and double_start <= today <= double_end:
            print(f"[LOG FASE] Fase ativa detectada: Duplado")
            print(f"[LOG FASE] Fase retornada: Duplado\n")
            return 'Duplado'

        # 2. Verificar condições especiais (Operação e Carregamento)
        if (field_op_date and today >= field_op_date and 
            self.course_status and ('concluído' in self.course_status.lower() or 'concluido' in self.course_status.lower())):
            if not loading_date or (loading_date and field_op_date >= loading_date):
                print(f"[LOG FASE] Fase especial: Operação")
                print(f"[LOG FASE] Fase retornada: Operação\n")
                return 'Operação'
            else:
                print(f"[LOG FASE] Condição não atendida para Operação: field_operation_date < loading_date")

        if loading_date and today >= loading_date:
            if double_end and loading_date >= double_end:
                print(f"[LOG FASE] Fase especial: Carregamento")
                print(f"[LOG FASE] Fase retornada: Carregamento\n")
                return 'Carregamento'

        # 3. Verificar se todas as fases estão no futuro
        all_phases_in_future = (
            (not integration_start or integration_start > today) and (not integration_end or integration_end > today) and
            (not normative_start or normative_start > today) and (not normative_end or normative_end > today) and
            (not technical_course_start or technical_course_start > today) and (not technical_course_end or technical_course_end > today) and
            (not double_start or double_start > today) and (not double_end or double_end > today)
        )
        has_any_phase = (
            (integration_start and integration_end) or
            (normative_start and normative_end) or
            (technical_course_start and technical_course_end) or
            (double_start and double_end)
        )
        if all_phases_in_future and has_any_phase:
            print("[LOG FASE] Todas as fases estão no futuro, retornando 'Previsto'")
            print(f"[LOG FASE] Fase retornada: Previsto\n")
            return 'Previsto'

        # 4. Verificar se há alguma fase futura
        has_future_phase = (
            (integration_start and integration_start > today) or
            (normative_start and normative_start > today) or
            (technical_course_start and technical_course_start > today) or
            (double_start and double_start > today)
        )
        if has_future_phase:
            print("[LOG FASE] Há fases futuras, retornando 'Previsto'")
            print(f"[LOG FASE] Fase retornada: Previsto\n")
            return 'Previsto'

        print("[LOG FASE] Nenhuma condição atendida, retornando 'Sem Fase Ativa'")
        print(f"[LOG FASE] Fase retornada: Sem Fase Ativa\n")
        return 'Sem Fase Ativa'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Routes
def safe_date_sort(a, b, field):
    """Função auxiliar para ordenação segura de datas, lidando com valores None"""
    a_val = getattr(a, field, None)
    b_val = getattr(b, field, None)
    
    if a_val is None and b_val is None:
        return 0
    if a_val is None:
        return 1  # Coloca None no final
    if b_val is None:
        return -1  # Coloca None no final
        
    if a_val < b_val:
        return -1
    elif a_val > b_val:
        return 1
    return 0




# Rotas de logout por marca
@app.route('/vivo/logout')
@login_required
def logout_vivo():
    logout_user()
    return redirect(url_for('login_vivo'))

@app.route('/claro/logout')
@login_required
def logout_claro():
    logout_user()
    return redirect(url_for('login_claro'))


# Rotas de gestão de usuários por marca
@app.route('/vivo/users')
@login_required
def user_management_vivo():
    if current_user.access_type != 'admin':
        return redirect(url_for('index_vivo'))
    users = User.query.filter_by(brand='Vivo').all()
    return render_template('users.html', users=users, brand='Vivo')

@app.route('/claro/users')
@login_required
def user_management_claro():
    if current_user.access_type != 'admin':
        return redirect(url_for('index_claro'))
    users = User.query.filter_by(brand='Claro').all()
    return render_template('users.html', users=users, brand='Claro')


# Rotas de cadastro de usuário por marca
@app.route('/vivo/add_user', methods=['GET', 'POST'])
@login_required
def add_user_vivo():
    if current_user.access_type != 'admin':
        return redirect(url_for('index_vivo'))
    if request.method == 'POST':
        username = request.form.get('username')
        password = generate_password_hash(request.form.get('password'), method='pbkdf2:sha256')
        name = request.form.get('name')
        access_type = request.form.get('access_type')
        if User.query.filter_by(username=username, brand='Vivo').first():
            flash('Usuário já existe')
            return redirect(url_for('add_user_vivo'))
        new_user = User(
            username=username,
            password=password,
            name=name,
            access_type=access_type,
            brand='Vivo'
        )
        db.session.add(new_user)
        db.session.commit()
        flash('Usuário adicionado com sucesso')
        return redirect(url_for('user_management_vivo'))
    return render_template('add_user.html', brand='Vivo')

@app.route('/claro/add_user', methods=['GET', 'POST'])
@login_required
def add_user_claro():
    if current_user.access_type != 'admin':
        return redirect(url_for('index_claro'))
    if request.method == 'POST':
        username = request.form.get('username')
        password = generate_password_hash(request.form.get('password'), method='pbkdf2:sha256')
        name = request.form.get('name')
        access_type = request.form.get('access_type')
        if User.query.filter_by(username=username, brand='Claro').first():
            flash('Usuário já existe')
            return redirect(url_for('add_user_claro'))
        new_user = User(
            username=username,
            password=password,
            name=name,
            access_type=access_type,
            brand='Claro'
        )
        db.session.add(new_user)
        db.session.commit()
        flash('Usuário adicionado com sucesso')
        return redirect(url_for('user_management_claro'))
    return render_template('add_user.html', brand='Claro')

@app.route('/vivo/edit_user/<int:user_id>', methods=['GET', 'POST'])
@login_required
def edit_user_vivo(user_id):
    if current_user.access_type != 'admin' and current_user.id != user_id:
        return redirect(url_for('index_vivo'))
    user = User.query.filter_by(id=user_id, brand='Vivo').first_or_404()
    if request.method == 'POST':
        user.name = request.form.get('name')
        if current_user.access_type == 'admin':
            user.access_type = request.form.get('access_type')
        if request.form.get('password'):
            user.password = generate_password_hash(request.form.get('password'), method='pbkdf2:sha256')
        db.session.commit()
        flash('Usuário atualizado com sucesso')
        return redirect(url_for('user_management_vivo'))
    return render_template('edit_user.html', user=user, brand='Vivo')

@app.route('/claro/edit_user/<int:user_id>', methods=['GET', 'POST'])
@login_required
def edit_user_claro(user_id):
    if current_user.access_type != 'admin' and current_user.id != user_id:
        return redirect(url_for('index_claro'))
    user = User.query.filter_by(id=user_id, brand='Claro').first_or_404()
    if request.method == 'POST':
        user.name = request.form.get('name')
        if current_user.access_type == 'admin':
            user.access_type = request.form.get('access_type')
        if request.form.get('password'):
            user.password = generate_password_hash(request.form.get('password'), method='pbkdf2:sha256')
        db.session.commit()
        flash('Usuário atualizado com sucesso')
        return redirect(url_for('user_management_claro'))
    return render_template('edit_user.html', user=user, brand='Claro')

@app.route('/vivo/delete_user/<int:user_id>')
@login_required
def delete_user_vivo(user_id):
    if current_user.access_type != 'admin':
        return redirect(url_for('index_vivo'))
    user = User.query.filter_by(id=user_id, brand='Vivo').first_or_404()
    if user.username == 'admin':
        flash('Não é possível excluir o usuário administrador')
        return redirect(url_for('user_management_vivo'))
    db.session.delete(user)
    db.session.commit()
    flash('Usuário excluído com sucesso')
    return redirect(url_for('user_management_vivo'))

@app.route('/claro/delete_user/<int:user_id>')
@login_required
def delete_user_claro(user_id):
    if current_user.access_type != 'admin':
        return redirect(url_for('index_claro'))
    user = User.query.filter_by(id=user_id, brand='Claro').first_or_404()
    if user.username == 'admin':
        flash('Não é possível excluir o usuário administrador')
        return redirect(url_for('user_management_claro'))
    db.session.delete(user)
    db.session.commit()
    flash('Usuário excluído com sucesso')
    return redirect(url_for('user_management_claro'))

def handle_upload_file(brand='Vivo'):
    """Função auxiliar para processar upload de arquivo"""
    if current_user.access_type != 'admin':
        return jsonify({'error': 'Acesso negado. Apenas administradores podem fazer upload de arquivos.'}), 403
        
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
    
    if file and file.filename.endswith(('.xlsx', '.xls')):
        try:
            # Lê o Excel ignorando a primeira linha (cabeçalho)
            df = pd.read_excel(file, header=None, skiprows=1)
            updated = 0
            created = 0
            
            def parse_date(date_val):
                if pd.isna(date_val) or date_val in ['', ' ', '  /  /    ']:
                    return None
                try:
                    if isinstance(date_val, str):
                        date_val = date_val.strip()
                        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d'):
                            try:
                                return datetime.strptime(date_val, fmt).date()
                            except ValueError:
                                continue
                        return None
                    elif hasattr(date_val, 'date'):
                        return date_val.date()
                    elif hasattr(date_val, 'strftime'):
                        return date_val
                    return None
                except Exception as e:
                    print(f"Erro ao converter data: {date_val} - {str(e)}")
                    return None
            
            # Índices das colunas (baseado na ordem do modelo)
            COLS = {
                'matricula': 0,
                'nome': 1,
                'funcao': 2,
                'tipo': 3,
                'data_admissao': 4,
                'cep': 5,
                'situacao': 6,
                'status_curso': 7,
                'turma': 8,
                'local_curso': 9,
                'gerente': 10,
                'gerente_corporativo': 11,
                'instrutor': 12,
                'contato': 13,
                'apto_operacao': 14,
                'inicio_integracao': 15,  # Início Integração
                'fim_integracao': 16,     # Término Integração
                'inicio_normativo': 17,    # Início Normativo
                'fim_normativo': 18,       # Término Normativo
                'inicio_curso_tecnico': 19, # Início Curso Técnico
                'fim_curso_tecnico': 20,    # Término Curso Técnico
                'inicio_duplado': 21,       # Início Duplado
                'fim_duplado': 22,          # Término Duplado
                'data_carregamento': 23,    # Data de Carregamento
                'data_operacao_campo': 24   # Data de Operação de Campo
            }
            
            for _, row in df.iterrows():
                try:
                    # Obter valores por posição
                    matricula = str(row[COLS['matricula']]) if pd.notna(row[COLS['matricula']]) else ''
                    
                    # Verificar se a linha está vazia
                    if not matricula.strip():
                        continue
                        
                    # Verificar se o colaborador já existe
                    employee = Employee.query.filter_by(registration=matricula).first()
                    
                    # Processar valores não nulos
                    def get_value(idx, default=''):
                        return row[idx] if pd.notna(row[idx]) else default
                        
                    if employee:
                        # Função auxiliar para atualizar campo e registrar mudança
                        def update_field(field_name, value, is_date=False):
                            old_value = getattr(employee, field_name)
                            new_value = parse_date(value) if is_date and value is not None else value
                            
                            if is_date and new_value is not None:
                                new_value = new_value.date() if hasattr(new_value, 'date') else new_value
                            
                            if old_value != new_value and new_value is not None:
                                setattr(employee, field_name, new_value)
                                log_change(
                                    registration=employee.registration,
                                    field_changed=field_name,
                                    old_value=old_value,
                                    new_value=new_value,
                                    change_source='upload'
                                )
                        
                        # Atualizar campos regulares
                        update_field('full_name', get_value(COLS['nome']))
                        update_field('role', get_value(COLS['funcao']))
                        update_field('employee_type', get_value(COLS['tipo']))
                        update_field('cep', get_value(COLS['cep']))
                        update_field('status', get_value(COLS['situacao']))
                        update_field('course_status', get_value(COLS['status_curso']))
                        update_field('team', get_value(COLS['turma']))
                        update_field('course_location', get_value(COLS['local_curso']))
                        update_field('manager', get_value(COLS['gerente']))
                        update_field('corporate_manager', get_value(COLS['gerente_corporativo']))
                        update_field('instructor', get_value(COLS['instrutor']))
                        update_field('contato', get_value(COLS['contato']))
                        update_field('operation_ready', get_value(COLS['apto_operacao']))
                        
                        # Atualizar campos de data
                        date_fields = {
                            'admission_date': row[COLS['data_admissao']],
                            'integration_start': row[COLS['inicio_integracao']],
                            'integration_end': row[COLS['fim_integracao']],
                            'normative_start': row[COLS['inicio_normativo']],
                            'normative_end': row[COLS['fim_normativo']],
                            'technical_course_start': row[COLS['inicio_curso_tecnico']],
                            'technical_course_end': row[COLS['fim_curso_tecnico']],
                            'double_start': row[COLS['inicio_duplado']],
                            'double_end': row[COLS['fim_duplado']],
                            'loading_date': row[COLS['data_carregamento']],
                            'field_operation_date': row[COLS['data_operacao_campo']]
                        }
                        
                        for field, value in date_fields.items():
                            if pd.notna(value):
                                update_field(field, value, is_date=True)
                        
                        employee.last_updated = get_brasil_time()
                        
                        updated += 1
                    else:
                        # Criar novo colaborador
                        new_employee = Employee(
                            registration=matricula,
                            brand=brand,
                            full_name=get_value(COLS['nome']),
                            role=get_value(COLS['funcao']),
                            employee_type=get_value(COLS['tipo']),
                            admission_date=parse_date(row[COLS['data_admissao']]),
                            cep=get_value(COLS['cep']),
                            status=get_value(COLS['situacao']),
                            course_status=get_value(COLS['status_curso']),
                            team=get_value(COLS['turma']),
                            course_location=get_value(COLS['local_curso']),
                            manager=get_value(COLS['gerente']),
                            corporate_manager=get_value(COLS['gerente_corporativo']),
                            instructor=get_value(COLS['instrutor']),
                            contato=get_value(COLS['contato']),
                            operation_ready=get_value(COLS['apto_operacao']),
                            integration_start=parse_date(row[COLS['inicio_integracao']]),
                            integration_end=parse_date(row[COLS['fim_integracao']]),
                            normative_start=parse_date(row[COLS['inicio_normativo']]),
                            normative_end=parse_date(row[COLS['fim_normativo']]),
                            technical_course_start=parse_date(row[COLS['inicio_curso_tecnico']]),
                            technical_course_end=parse_date(row[COLS['fim_curso_tecnico']]),
                            double_start=parse_date(row[COLS['inicio_duplado']]),
                            double_end=parse_date(row[COLS['fim_duplado']]),
                            loading_date=parse_date(row[COLS['data_carregamento']]),
                            field_operation_date=parse_date(row[COLS['data_operacao_campo']]),
                            last_updated=get_brasil_time()
                        )
                        
                        # Registrar a criação do novo colaborador
                        log_change(
                            registration=matricula,
                            field_changed='new_employee',
                            old_value=None,
                            new_value='Novo colaborador criado',
                            change_source='upload'
                        )
                        db.session.add(new_employee)
                        created += 1
                        
                except Exception as e:
                    print(f"Erro ao processar linha: {row}\nErro: {str(e)}")
                    continue
            
            db.session.commit()
            return jsonify({
                'message': f'Importação concluída com sucesso! {created} registros criados, {updated} registros atualizados.'
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Erro ao processar o arquivo: {str(e)}'}), 500
    
    return jsonify({'error': 'Formato de arquivo não suportado'}), 400

@app.route('/vivo/upload', methods=['POST'])
@login_required
def upload_file_vivo():
    return handle_upload_file(brand='Vivo')

@app.route('/claro/upload', methods=['POST'])
@login_required
def upload_file_claro():
    return handle_upload_file(brand='Claro')

@app.route('/vivo/employee/view/<int:employee_id>')
@app.route('/vivo/employee/view/<int:employee_id>/<referrer>')
@login_required
def view_employee_vivo(employee_id, referrer=None):
    employee = Employee.query.filter_by(id=employee_id, brand='Vivo').first_or_404()
    return render_template('view_employee.html', employee=employee, now=datetime.now(), referrer=referrer, brand='Vivo')

@app.route('/claro/employee/view/<int:employee_id>')
@app.route('/claro/employee/view/<int:employee_id>/<referrer>')
@login_required
def view_employee_claro(employee_id, referrer=None):
    employee = Employee.query.filter_by(id=employee_id, brand='Claro').first_or_404()
    return render_template('view_employee.html', employee=employee, now=datetime.now(), referrer=referrer, brand='Claro')

@app.route('/employee/view/<int:employee_id>')
@app.route('/employee/view/<int:employee_id>/<referrer>')
@login_required
def view_employee(employee_id, referrer=None):
    employee = Employee.query.get_or_404(employee_id)
    return render_template('view_employee.html', employee=employee, now=datetime.now(), referrer=referrer)

@app.route('/vivo/employee/edit/<int:employee_id>', methods=['GET', 'POST'])
@login_required
def edit_employee_vivo(employee_id):
    if current_user.access_type != 'admin':
        flash('Acesso negado. Apenas administradores podem editar colaboradores.', 'danger')
        return redirect(url_for('view_employee_vivo', employee_id=employee_id))
    employee = Employee.query.filter_by(id=employee_id, brand='Vivo').first_or_404()
    if request.method == 'POST':
        employee.full_name = request.form.get('full_name', employee.full_name)
        employee.contato = request.form.get('contato', employee.contato)
        db.session.commit()
        flash('Dados do colaborador atualizados com sucesso', 'success')
        return redirect(url_for('view_employee_vivo', employee_id=employee_id))
    return render_template('edit_employee.html', employee=employee, brand='Vivo')

@app.route('/claro/employee/edit/<int:employee_id>', methods=['GET', 'POST'])
@login_required
def edit_employee_claro(employee_id):
    if current_user.access_type != 'admin':
        flash('Acesso negado. Apenas administradores podem editar colaboradores.', 'danger')
        return redirect(url_for('view_employee_claro', employee_id=employee_id))
    employee = Employee.query.filter_by(id=employee_id, brand='Claro').first_or_404()
    if request.method == 'POST':
        employee.full_name = request.form.get('full_name', employee.full_name)
        employee.contato = request.form.get('contato', employee.contato)
        db.session.commit()
        flash('Dados do colaborador atualizados com sucesso', 'success')
        return redirect(url_for('view_employee_claro', employee_id=employee_id))
    return render_template('edit_employee.html', employee=employee, brand='Claro')

@app.route('/employee/edit/<int:employee_id>', methods=['GET', 'POST'])
@login_required
def edit_employee(employee_id):
    if current_user.access_type != 'admin':
        flash('Acesso negado. Apenas administradores podem editar colaboradores.', 'danger')
        return redirect(url_for('view_employee', employee_id=employee_id))
        
    employee = Employee.query.get_or_404(employee_id)
    
    if request.method == 'POST':
        # Atualiza o nome e o contato
        employee.full_name = request.form.get('full_name', employee.full_name)
        employee.contato = request.form.get('contato', employee.contato)
        db.session.commit()
        flash('Dados do colaborador atualizados com sucesso', 'success')
        return redirect(url_for('view_employee', employee_id=employee_id))
    
    return render_template('edit_employee.html', employee=employee)

@app.route('/vivo/api/employee/<int:employee_id>', methods=['PUT'])
@login_required
def update_employee_api_vivo(employee_id):
    if current_user.access_type != 'admin':
        return jsonify({'error': 'Acesso negado. Apenas administradores podem editar colaboradores.'}), 403
    employee = Employee.query.filter_by(id=employee_id, brand='Vivo').first_or_404()
    data = request.get_json()
    for field in ['full_name', 'role', 'employee_type', 'status', 'course_status', 'team', 
                 'course_location', 'manager', 'corporate_manager', 'instructor', 'contato', 'operation_ready']:
        if field in data and getattr(employee, field) != data[field]:
            old_value = getattr(employee, field)
            setattr(employee, field, data[field])
            log_change(registration=employee.registration, field_changed=field, old_value=old_value, new_value=data[field], change_source='system')
    date_fields = {'admission_date': '%Y-%m-%d', 'integration_start': '%Y-%m-%d', 'integration_end': '%Y-%m-%d',
                   'normative_start': '%Y-%m-%d', 'normative_end': '%Y-%m-%d', 'technical_course_start': '%Y-%m-%d',
                   'technical_course_end': '%Y-%m-%d', 'double_start': '%Y-%m-%d', 'double_end': '%Y-%m-%d',
                   'loading_date': '%Y-%m-%d', 'field_operation_date': '%Y-%m-%d'}
    for field, fmt in date_fields.items():
        if field in data:
            try:
                new_date = datetime.strptime(data[field], fmt).date() if data[field] else None
                old_date = getattr(employee, field)
                if old_date != new_date:
                    setattr(employee, field, new_date)
                    log_change(registration=employee.registration, field_changed=field, old_value=old_date, new_value=new_date, change_source='system')
            except ValueError:
                return jsonify({'error': f'Formato de data inválido para o campo {field}'}), 400
    db.session.commit()
    return jsonify({'message': 'Dados atualizados com sucesso', 'employee': {'id': employee.id, 'full_name': employee.full_name, 'registration': employee.registration, 'admission_date': employee.admission_date.strftime('%d/%m/%Y') if employee.admission_date else None, 'current_phase': employee.get_current_phase()}})

@app.route('/claro/api/employee/<int:employee_id>', methods=['PUT'])
@login_required
def update_employee_api_claro(employee_id):
    if current_user.access_type != 'admin':
        return jsonify({'error': 'Acesso negado. Apenas administradores podem editar colaboradores.'}), 403
    employee = Employee.query.filter_by(id=employee_id, brand='Claro').first_or_404()
    data = request.get_json()
    for field in ['full_name', 'role', 'employee_type', 'status', 'course_status', 'team', 
                 'course_location', 'manager', 'corporate_manager', 'instructor', 'contato', 'operation_ready']:
        if field in data and getattr(employee, field) != data[field]:
            old_value = getattr(employee, field)
            setattr(employee, field, data[field])
            log_change(registration=employee.registration, field_changed=field, old_value=old_value, new_value=data[field], change_source='system')
    date_fields = {'admission_date': '%Y-%m-%d', 'integration_start': '%Y-%m-%d', 'integration_end': '%Y-%m-%d',
                   'normative_start': '%Y-%m-%d', 'normative_end': '%Y-%m-%d', 'technical_course_start': '%Y-%m-%d',
                   'technical_course_end': '%Y-%m-%d', 'double_start': '%Y-%m-%d', 'double_end': '%Y-%m-%d',
                   'loading_date': '%Y-%m-%d', 'field_operation_date': '%Y-%m-%d'}
    for field, fmt in date_fields.items():
        if field in data:
            try:
                new_date = datetime.strptime(data[field], fmt).date() if data[field] else None
                old_date = getattr(employee, field)
                if old_date != new_date:
                    setattr(employee, field, new_date)
                    log_change(registration=employee.registration, field_changed=field, old_value=old_date, new_value=new_date, change_source='system')
            except ValueError:
                return jsonify({'error': f'Formato de data inválido para o campo {field}'}), 400
    db.session.commit()
    return jsonify({'message': 'Dados atualizados com sucesso', 'employee': {'id': employee.id, 'full_name': employee.full_name, 'registration': employee.registration, 'admission_date': employee.admission_date.strftime('%d/%m/%Y') if employee.admission_date else None, 'current_phase': employee.get_current_phase()}})

@app.route('/api/employee/<int:employee_id>', methods=['PUT'])
@login_required
def update_employee_api(employee_id):
    if current_user.access_type != 'admin':
        return jsonify({'error': 'Acesso negado. Apenas administradores podem editar colaboradores.'}), 403
        
    employee = Employee.query.get_or_404(employee_id)
    data = request.get_json()
    
    # Update only the fields that are provided in the request
    for field in ['full_name', 'role', 'employee_type', 'status', 'course_status', 'team', 
                 'course_location', 'manager', 'corporate_manager', 'instructor', 'contato', 'operation_ready']:
        if field in data and getattr(employee, field) != data[field]:
            old_value = getattr(employee, field)
            setattr(employee, field, data[field])
            log_change(
                registration=employee.registration,
                field_changed=field,
                old_value=old_value,
                new_value=data[field],
                change_source='system'
            )
    
    # Handle date fields
    date_fields = {
        'admission_date': '%Y-%m-%d',
        'integration_start': '%Y-%m-%d',
        'integration_end': '%Y-%m-%d',
        'normative_start': '%Y-%m-%d',
        'normative_end': '%Y-%m-%d',
        'technical_course_start': '%Y-%m-%d',
        'technical_course_end': '%Y-%m-%d',
        'double_start': '%Y-%m-%d',
        'double_end': '%Y-%m-%d',
        'loading_date': '%Y-%m-%d',
        'field_operation_date': '%Y-%m-%d'
    }
    
    for field, fmt in date_fields.items():
        if field in data:
            try:
                new_date = datetime.strptime(data[field], fmt).date() if data[field] else None
                old_date = getattr(employee, field)
                
                if old_date != new_date:
                    setattr(employee, field, new_date)
                    log_change(
                        registration=employee.registration,
                        field_changed=field,
                        old_value=old_date,
                        new_value=new_date,
                        change_source='system'
                    )
            except ValueError:
                return jsonify({'error': f'Formato de data inválido para o campo {field}'}), 400
    
    db.session.commit()
    return jsonify({
        'message': 'Dados atualizados com sucesso',
        'employee': {
            'id': employee.id,
            'full_name': employee.full_name,
            'registration': employee.registration,
            'admission_date': employee.admission_date.strftime('%d/%m/%Y') if employee.admission_date else None,
            'current_phase': employee.get_current_phase()
        }
    })

@app.route('/vivo/delete_employee/<int:employee_id>', methods=['POST'])
@login_required
def delete_employee_vivo(employee_id):
    if current_user.access_type != 'admin':
        flash('Acesso negado. Apenas administradores podem excluir colaboradores.', 'danger')
        return redirect(url_for('index_vivo'))
    if request.method == 'POST':
        employee = Employee.query.filter_by(id=employee_id, brand='Vivo').first_or_404()
        db.session.delete(employee)
        db.session.commit()
        return jsonify({'message': 'Colaborador excluído com sucesso'})
    return jsonify({'error': 'Método não permitido'}), 405

@app.route('/claro/delete_employee/<int:employee_id>', methods=['POST'])
@login_required
def delete_employee_claro(employee_id):
    if current_user.access_type != 'admin':
        flash('Acesso negado. Apenas administradores podem excluir colaboradores.', 'danger')
        return redirect(url_for('index_claro'))
    if request.method == 'POST':
        employee = Employee.query.filter_by(id=employee_id, brand='Claro').first_or_404()
        db.session.delete(employee)
        db.session.commit()
        return jsonify({'message': 'Colaborador excluído com sucesso'})
    return jsonify({'error': 'Método não permitido'}), 405

@app.route('/delete_employee/<int:employee_id>', methods=['POST'])
@login_required
def delete_employee(employee_id):
    if current_user.access_type != 'admin':
        flash('Acesso negado. Apenas administradores podem excluir colaboradores.', 'danger')
        return redirect(url_for('index'))
        
    if request.method == 'POST':
        employee = Employee.query.get_or_404(employee_id)
        db.session.delete(employee)
        db.session.commit()
        return jsonify({'message': 'Colaborador excluído com sucesso'})
    return jsonify({'error': 'Método não permitido'}), 405

@app.route('/vivo/delete_employees', methods=['POST'])
@login_required
def delete_employees_vivo():
    if current_user.access_type != 'admin':
        return jsonify({'error': 'Acesso negado. Apenas administradores podem excluir colaboradores.'}), 403
    if request.method == 'POST':
        try:
            data = request.get_json()
            employee_ids = data.get('employeeIds', [])
            if not employee_ids:
                return jsonify({'error': 'Nenhum colaborador selecionado'}), 400
            for i in range(0, len(employee_ids), 100):
                batch = employee_ids[i:i+100]
                Employee.query.filter(Employee.id.in_(batch), Employee.brand == 'Vivo').delete(synchronize_session=False)
                db.session.commit()
            return jsonify({'message': f'{len(employee_ids)} colaborador(es) excluído(s) com sucesso'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Erro ao excluir colaboradores: {str(e)}'}), 500
    return jsonify({'error': 'Método não permitido'}), 405

@app.route('/claro/delete_employees', methods=['POST'])
@login_required
def delete_employees_claro():
    if current_user.access_type != 'admin':
        return jsonify({'error': 'Acesso negado. Apenas administradores podem excluir colaboradores.'}), 403
    if request.method == 'POST':
        try:
            data = request.get_json()
            employee_ids = data.get('employeeIds', [])
            if not employee_ids:
                return jsonify({'error': 'Nenhum colaborador selecionado'}), 400
            for i in range(0, len(employee_ids), 100):
                batch = employee_ids[i:i+100]
                Employee.query.filter(Employee.id.in_(batch), Employee.brand == 'Claro').delete(synchronize_session=False)
                db.session.commit()
            return jsonify({'message': f'{len(employee_ids)} colaborador(es) excluído(s) com sucesso'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Erro ao excluir colaboradores: {str(e)}'}), 500
    return jsonify({'error': 'Método não permitido'}), 405

@app.route('/delete_employees', methods=['POST'])
@login_required
def delete_employees():
    if current_user.access_type != 'admin':
        return jsonify({'error': 'Acesso negado. Apenas administradores podem excluir colaboradores.'}), 403
        
    if request.method == 'POST':
        try:
            data = request.get_json()
            employee_ids = data.get('employeeIds', [])
            
            if not employee_ids:
                return jsonify({'error': 'Nenhum colaborador selecionado'}), 400
            
            # Delete employees in batches to avoid large transactions
            for i in range(0, len(employee_ids), 100):
                batch = employee_ids[i:i+100]
                Employee.query.filter(Employee.id.in_(batch)).delete(synchronize_session=False)
                db.session.commit()
            
            return jsonify({'message': f'{len(employee_ids)} colaborador(es) excluído(s) com sucesso'})
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Erro ao excluir colaboradores: {str(e)}'}), 500
    
    return jsonify({'error': 'Método não permitido'}), 405

# Create admin user if not exists
def create_admin_user():
    # Criar admin para Vivo
    if not User.query.filter_by(username='RafaelPinho', brand='Vivo').first():
        admin_vivo = User(
            username='RafaelPinho',
            password=generate_password_hash('@21314100', method='pbkdf2:sha256'),
            name='Administrador Vivo',
            access_type='admin',
            brand='Vivo'
        )
        db.session.add(admin_vivo)
    
    # Criar admin para Claro
    if not User.query.filter_by(username='RafaelPinho', brand='Claro').first():
        admin_claro = User(
            username='RafaelPinho',
            password=generate_password_hash('@21314100', method='pbkdf2:sha256'),
            name='Administrador Claro',
            access_type='admin',
            brand='Claro'
        )
        db.session.add(admin_claro)
    
    db.session.commit()

@app.route('/vivo/download_modelo')
@login_required
def download_modelo_vivo():
    return download_modelo_impl()

@app.route('/claro/download_modelo')
@login_required
def download_modelo_claro():
    return download_modelo_impl()

def download_modelo_impl():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    import io
    
    # Criar um novo workbook e selecionar a planilha ativa
    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo de Importação"
    
    # Criar planilha de instruções
    ws_instrucoes = wb.create_sheet('Instruções', 0)
    ws_instrucoes['A1'] = 'INSTRUÇÕES DE USO:'
    ws_instrucoes['A1'].font = Font(bold=True, size=14)
    
    # Adicionar instruções
    instrucoes = [
        ('1.', 'Este é o modelo para importação de colaboradores.'),
        ('2.', 'A planilha "Modelo de Importação" contém as colunas necessárias.'),
        ('3.', 'Formato de datas: DD/MM/AAAA'),
        ('4.', 'TODOS OS CAMPOS DEVEM ESTAR SEM ACENTOS E EM MINÚSCULOS'),
        ('5.', 'O sistema aceita variações dos nomes (com/sem "do", "da", etc)'),
        ('6.', 'Exemplos de nomes alternativos aceitos:'),
        ('', '  - "nome" ou "nome completo"'),
        ('', '  - "status" ou "status curso" ou "status do curso tecnico"'),
        ('', '  - "telefone" ou "contato"'),
        ('', '  - "inicio integracao" ou "inicio da integracao"'),
        ('', '  - "termino integracao" ou "termino da integracao"'),
        ('', '  - "inicio normativo" ou "inicio do normativo"'),
        ('', '  - "inicio curso" ou "inicio do curso"'),
        ('', '  - "data op campo" ou "data operacao campo"')
    ]
    
    # Aplicar formatação às instruções
    for i, (num, texto) in enumerate(instrucoes, start=1):
        if num:  # Se tiver número (não for linha de continuação)
            ws_instrucoes[f'A{i+1}'] = num
            ws_instrucoes[f'B{i+1}'] = texto
        else:  # Se for continuação
            ws_instrucoes[f'B{i+1}'] = texto
    
    # Ajustar largura das colunas da planilha de instruções
    ws_instrucoes.column_dimensions['A'].width = 5
    ws_instrucoes.column_dimensions['B'].width = 60
    
    # Criar a planilha de modelo
    ws = wb['Modelo de Importação']
    
    # Definir os cabeçalhos padronizados (sem acentos e em minúsculas)
    headers = [
        'matricula', 'nome completo', 'funcao', 'tipo', 'data admissao',
        'cep', 'situacao', 'status curso', 'turma', 'local curso',
        'gerente', 'gerente corporativo', 'instrutor', 'contato', 'apto operacao',
        'inicio integracao', 'termino integracao',
        'inicio normativo', 'termino normativo',
        'inicio curso tecnico', 'termino curso tecnico',
        'inicio duplado', 'termino duplado',
        'data carregamento', 'data operacao campo'
    ]
    
    # Adicionar cabeçalhos em negrito
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Adicionar exemplo de dados
    example_data = [
        '12345', 'Fulano de Tal', 'Operador', 'CLT', '01/01/2023',
        '12345-678', 'Ativo', 'Concluído', 'Turma A', 'São Paulo',
        'João Silva', 'Maria Oliveira', 'Carlos Santos', '(11) 98765-4321', 'Sim',
        '01/02/2023', '28/02/2023', '01/03/2023', '31/03/2023',
        '01/04/2023', '30/04/2023', '01/05/2023', '31/05/2023',
        '01/06/2023', '15/06/2023'
    ]
    
    # Adicionar uma linha explicando que pode usar os nomes alternativos
    ws['A3'] = 'OBS: Os nomes entre parênteses são alternativos que também são aceitos pelo sistema.'
    ws.merge_cells('A3:Y3')
    ws['A3'].font = Font(italic=True, color='FF0000')
    ws['A3'].alignment = Alignment(horizontal='center')
    
    for col_num, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col_num, value=value)
        # Formatar células de data
        if col_num in [5, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25]:
            cell.number_format = 'DD/MM/YYYY'
    
    # Ajustar a largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = min(adjusted_width, 30)
    
    # Criar um buffer para o arquivo
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Retornar o arquivo
    return send_file(
        buffer,
        as_attachment=True,
        download_name='modelo_importacao_colaboradores.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/relatorio/gerentes')
@login_required
def relatorio_gerentes():
    from datetime import datetime
    # Determinar a marca do usuário (se autenticado) para filtrar dados
    brand = getattr(current_user, 'brand', None) if current_user and hasattr(current_user, 'is_authenticated') and current_user.is_authenticated else None

    # Obter todos os gerentes corporativos distintos (filtrando pela marca quando disponível)
    gerentes_q = db.session.query(Employee.corporate_manager).distinct().filter(
        Employee.corporate_manager.isnot(None)
    )
    if brand:
        gerentes_q = gerentes_q.filter(Employee.brand == brand)
    gerentes = gerentes_q.order_by(Employee.corporate_manager).all()
    gerentes = [g[0] for g in gerentes if g[0]]  # Remover valores nulos
    
    # Obter todos os tipos distintos
    tipos_q = db.session.query(Employee.employee_type).distinct().filter(
        Employee.employee_type.isnot(None)
    )
    if brand:
        tipos_q = tipos_q.filter(Employee.brand == brand)
    tipos = tipos_q.order_by(Employee.employee_type).all()
    tipos = [t[0] for t in tipos if t[0]]  # Remover valores nulos
    
    # Obter todas as datas de operação distintas
    datas_q = db.session.query(
        Employee.field_operation_date,
        Employee.employee_type
    ).distinct().filter(
        Employee.field_operation_date.isnot(None),
        Employee.employee_type.isnot(None)
    )
    if brand:
        datas_q = datas_q.filter(Employee.brand == brand)
    datas_operacao = datas_q.order_by(Employee.field_operation_date, Employee.employee_type).all()
    
    # Criar estrutura de dados para datas e tipos
    datas_tipos = {}
    fases_por_data = {}
    for data, tipo in datas_operacao:
        if data:
            data_str = data.strftime('%d/%m/%Y')
            if data_str not in datas_tipos:
                datas_tipos[data_str] = []
                # Obter fases de todos os colaboradores para esta data
                fases_q = db.session.query(Employee).filter(
                    Employee.field_operation_date == data
                )
                if brand:
                    fases_q = fases_q.filter(Employee.brand == brand)
                fases = fases_q.all()
                fases_unicas = set()
                for emp in fases:
                    fase = emp.get_current_phase()
                    print(f"[LOG RELATORIO] Data: {data_str} | Matrícula: {emp.registration} | Fase: {fase}")
                    if fase and fase != 'Sem fase ativa':
                        fases_unicas.add(fase)

                # Escolher fase preferencial por prioridade (evita seleção por ordenação alfabética)
                prioridade = ['Integração', 'Normativo', 'Curso Técnico', 'Duplado', 'Operação', 'Carregamento', 'Previsto']
                preferred = None
                for p in prioridade:
                    if p in fases_unicas:
                        preferred = p
                        break

                if preferred:
                    fases_por_data[data_str] = [preferred]
                else:
                    # fallback: lista ordenada (vazia se não houver)
                    print(f"[LOG RELATORIO] Data: {data_str} | Fases únicas: {sorted(list(fases_unicas)) if fases_unicas else []}")
                    fases_por_data[data_str] = sorted(list(fases_unicas)) if fases_unicas else []
            if tipo and tipo not in datas_tipos[data_str]:
                datas_tipos[data_str].append(tipo)
    
    # Criar dicionário com a contagem e datas de admissão por gerente, data e tipo
    dados = {}
    for gerente in gerentes:
        dados[gerente] = {}
        for data_str, tipos_data in datas_tipos.items():
            data = datetime.strptime(data_str, '%d/%m/%Y').date()
            dados[gerente][data_str] = {}
            for tipo in tipos_data:
                # Buscar todos os colaboradores para contagem total
                colaboradores_q = Employee.query.filter(
                    Employee.corporate_manager == gerente,
                    Employee.field_operation_date == data,
                    Employee.employee_type == tipo
                )
                if brand:
                    colaboradores_q = colaboradores_q.filter(Employee.brand == brand)
                colaboradores_todos = colaboradores_q.all()
                
                # Buscar apenas os colaboradores aptos (SIM) para a tabela principal
                colaboradores_q2 = Employee.query.filter(
                    Employee.corporate_manager == gerente,
                    Employee.field_operation_date == data,
                    Employee.employee_type == tipo,
                    db.or_(
                        db.func.lower(Employee.operation_ready) == 'sim',
                        db.func.upper(Employee.operation_ready) == 'SIM',
                        db.func.lower(Employee.operation_ready) == 'sin',
                        db.func.upper(Employee.operation_ready) == 'SIN',
                        db.func.lower(Employee.operation_ready).like('sim%'),
                        db.func.lower(Employee.operation_ready).like('sin%')
                    )
                )
                if brand:
                    colaboradores_q2 = colaboradores_q2.filter(Employee.brand == brand)
                colaboradores = colaboradores_q2.all()
                
                # Usar todos os colaboradores para a contagem total
                colaboradores_info = []
                for col in colaboradores_todos:
                    print(f"[LOG RELATORIO] Gerente: {gerente} | Data: {data_str} | Tipo: {tipo} | Matrícula: {col.registration} | Fase: {col.get_current_phase()}")
                    colaboradores_info.append({
                        'operation_ready': col.operation_ready,
                        'course_status': col.course_status,
                        'admission_date': col.admission_date.strftime('%d/%m/%Y') if col.admission_date else None,
                        'registration': col.registration,
                        'full_name': col.full_name
                    })
                
                # Usar apenas os aptos para a contagem da tabela principal
                contagem_aptos = len(colaboradores)
                contagem_total = len(colaboradores_todos)
                
                if contagem_total > 0:
                    # Formatar as datas de admissão dos aptos
                    datas_admissao = []
                    for col in colaboradores:  # Usando apenas os aptos para as datas de admissão
                        if col.admission_date:
                            try:
                                data_formatada = col.admission_date.strftime('%d/%m/%Y')
                                datas_admissao.append(data_formatada)
                            except:
                                # Se houver erro ao formatar, usar a data como string
                                datas_admissao.append(str(col.admission_date))
                    
                    # Usar a contagem de aptos para a célula principal
                    # mas manter todos os colaboradores (aptos e não aptos) na lista de colaboradores
                    dados[gerente][data_str][tipo] = {
                        'quantidade': contagem_aptos,  # Mostrar apenas a contagem de aptos
                        'datas_admissao': datas_admissao if datas_admissao else [],
                        'colaboradores': colaboradores_info  # Mantendo todos para a linha de não aptos
                    }
    
    # Função para formatar tooltip com as datas de admissão
    def formatar_tooltip(datas):
        if not datas:
            return ""
        if len(datas) == 1:
            return f"Admitido em: {datas[0]}"
        return "Admissões: " + ", ".join(datas)

    # Calcular o total de colaboradores
    total_colaboradores = 0
    for gerente_data in dados.values():
        for data_tipos in gerente_data.values():
            for tipo_info in data_tipos.values():
                total_colaboradores += tipo_info['quantidade']

    # Ordenar as datas
    from datetime import datetime
    datas_ordenadas = sorted(datas_tipos.keys(), key=lambda x: datetime.strptime(x, '%d/%m/%Y'))
    datas_tipos_ordenado = {data: datas_tipos[data] for data in datas_ordenadas}
    
    # Obter a data da última atualização diretamente do modelo Employee
    from sqlalchemy import desc
    
    # Consulta para obter o registro mais recente (filtrando por marca quando aplicável)
    ultimo_q = Employee.query
    if brand:
        ultimo_q = ultimo_q.filter(Employee.brand == brand)
    ultimo_registro = ultimo_q.order_by(desc(Employee.last_updated)).first()
    
    if ultimo_registro and ultimo_registro.last_updated:
        ultima_atualizacao = ultimo_registro.last_updated
        print(f"Data mais recente encontrada: {ultima_atualizacao}")
    else:
        ultima_atualizacao = datetime.utcnow()
        print("Usando data atual como última atualização")
    
    # Garantir que temos um objeto datetime
    if not isinstance(ultima_atualizacao, datetime):
        print(f"Aviso: Formato de data inesperado: {ultima_atualizacao} (tipo: {type(ultima_atualizacao)})")
        ultima_atualizacao = datetime.utcnow()
    
    # Verificar se a data já está em UTC (sem informação de timezone)
    # Se estiver em UTC, converter para o fuso horário de Brasília (UTC-3)
    if ultima_atualizacao.tzinfo is None:
        # Assumir que está em UTC e converter para o fuso horário de Brasília (UTC-3)
        from datetime import timezone, timedelta
        brasil_tz = timezone(timedelta(hours=-3))
        ultima_atualizacao = ultima_atualizacao.astimezone(brasil_tz)
    else:
        # Se já tiver timezone, apenas converter para o fuso de Brasília
        from datetime import timezone, timedelta
        brasil_tz = timezone(timedelta(hours=-3))
        ultima_atualizacao = ultima_atualizacao.astimezone(brasil_tz)
    
    print(f"Data após conversão para Brasília: {ultima_atualizacao}")
    
    # Forçar o fechamento da sessão atual para evitar cache
    db.session.remove()

    return render_template('relatorio_gerentes.html', 
                         gerentes=gerentes, 
                         datas_tipos=datas_tipos_ordenado,
                         fases_por_data=fases_por_data,
                         ultima_atualizacao=ultima_atualizacao,
                         dados=dados,
                         total_colaboradores=total_colaboradores,
                         formatar_tooltip=formatar_tooltip,
                         tipos=tipos,
                         now=datetime.now)

@app.route('/dashboard_fases')
@login_required
def dashboard_fases():
    # Obter parâmetros de filtro
    apto_operacao = request.args.get('apto_operacao', 'todos')
    mes_operacao = request.args.get('mes_operacao', 'todos')
    
    # Determinar a marca do usuário (se autenticado) para filtrar dados
    brand = getattr(current_user, 'brand', None) if current_user and hasattr(current_user, 'is_authenticated') and current_user.is_authenticated else None

    # Obter lista de meses/anos disponíveis no banco de dados (filtrando por marca quando aplicável)
    meses_q = db.session.query(
        db.func.strftime('%Y-%m', Employee.field_operation_date).label('mes_ano')
    ).filter(
        Employee.field_operation_date.isnot(None)
    )
    if brand:
        meses_q = meses_q.filter(Employee.brand == brand)
    meses_disponiveis = meses_q.distinct().order_by(
        db.desc('mes_ano')
    ).all()
    
    # Formatar a lista de meses disponíveis para o template
    meses_formatados = []
    for mes_ano in meses_disponiveis:
        if mes_ano[0]:  # Verifica se não é None
            try:
                # Converte de 'YYYY-MM' para (valor, label) ex: ('2023-11', 'Novembro/2023')
                ano, mes = mes_ano[0].split('-')
                mes_nome = [
                    'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                    'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'
                ][int(mes)-1]
                meses_formatados.append((mes_ano[0], f"{mes_nome}/{ano}"))
            except (ValueError, IndexError):
                continue
    
    # Obter todos os colaboradores com filtros opcionais
    query = Employee.query
    if brand:
        query = query.filter(Employee.brand == brand)
    
    # Aplicar filtro de mês de operação
    if mes_operacao and mes_operacao != 'todos':
        try:
            mes = int(mes_operacao)
            if 1 <= mes <= 12:
                query = query.filter(
                    db.func.extract('month', Employee.field_operation_date) == mes
                )
        except (ValueError, TypeError):
            pass  # Se houver erro no parâmetro, ignora o filtro
    
    # Aplicar filtro de apto para operação
    if apto_operacao == 'sim':
        # Busca por 'Sim' em qualquer formato (case insensitive)
        query = query.filter(
            db.or_(
                db.func.lower(Employee.operation_ready) == 'sim',
                db.func.lower(Employee.operation_ready) == 's',
                db.func.lower(Employee.operation_ready) == 'y',
                db.func.lower(Employee.operation_ready) == 'yes',
                db.func.lower(Employee.operation_ready) == '1',
                Employee.operation_ready == 'Sim',
                Employee.operation_ready == 'SIM',
                Employee.operation_ready == 'S',
                Employee.operation_ready == 's',
                Employee.operation_ready == '1'
            )
        )
    elif apto_operacao == 'nao':
        # Busca por qualquer coisa que não seja considerado 'Sim'
        query = query.filter(
            db.or_(
                Employee.operation_ready.is_(None),
                Employee.operation_ready == '',
                db.func.lower(Employee.operation_ready) == 'não',
                db.func.lower(Employee.operation_ready) == 'nao',
                db.func.lower(Employee.operation_ready) == 'n',
                db.func.lower(Employee.operation_ready) == 'no',
                db.func.lower(Employee.operation_ready) == '0',
                Employee.operation_ready == 'Não',
                Employee.operation_ready == 'NÃO',
                Employee.operation_ready == 'N',
                Employee.operation_ready == 'n',
                Employee.operation_ready == '0',
                # Inclui qualquer outro valor que não seja considerado 'Sim'
                ~db.or_(
                    db.func.lower(Employee.operation_ready) == 'sim',
                    db.func.lower(Employee.operation_ready) == 's',
                    db.func.lower(Employee.operation_ready) == 'y',
                    db.func.lower(Employee.operation_ready) == 'yes',
                    db.func.lower(Employee.operation_ready) == '1',
                    Employee.operation_ready == 'Sim',
                    Employee.operation_ready == 'SIM',
                    Employee.operation_ready == 'S',
                    Employee.operation_ready == 's',
                    Employee.operation_ready == '1'
                )
            )
        )
    
    employees = query.order_by(Employee.full_name).all()
    
    # Contar colaboradores por fase (apenas fases ativas)
    # As chaves devem corresponder exatamente ao retorno de get_current_phase
    phase_counts = {
        'Integração': 0,
        'Normativo': 0,
        'Curso Técnico': 0,
        'Duplado': 0,
        'Carregamento': 0,
        'Operação': 0,
        'Sem Fase Ativa': 0,
        'Previsto': 0
    }
    
    # Contar colaboradores em cada fase
    for employee in employees:
        phase = employee.get_current_phase()
        # Contar apenas as fases que estão na lista phase_counts
        if phase in phase_counts:
            phase_counts[phase] += 1
    
    # Preparar dados para o gráfico
    labels = list(phase_counts.keys())
    data = list(phase_counts.values())
    
    # Obter lista de colaboradores por fase (apenas fases ativas)
    employees_by_phase = {}
    for phase in phase_counts.keys():
        employees_by_phase[phase] = []
    
    for employee in employees:
        phase = employee.get_current_phase()
        # Incluir apenas se a fase estiver na lista de fases ativas
        if phase in employees_by_phase:
            employees_by_phase[phase].append(employee)
    
    # Ordenar colaboradores por data de operação de campo (mais recente primeiro)
    def to_date(dt):
        if dt is None:
            return None
        if isinstance(dt, datetime):
            return dt.date()
        return dt
    
    for phase in employees_by_phase:
        employees_by_phase[phase].sort(
            key=lambda x: to_date(x.field_operation_date) if x.field_operation_date else datetime.min.date(),
            reverse=True
        )
    
    return render_template(
        'dashboard_fases.html',
        phase_counts=phase_counts,
        labels=labels,
        data=data,
        employees_by_phase=employees_by_phase,
        now=datetime.now().date(),
        filtro_apto_operacao=apto_operacao,
        mes_selecionado=mes_operacao,
        meses_disponiveis=meses_formatados
    )

@app.route('/vivo/export_employees_excel')
@login_required
def export_employees_excel_vivo():
    return export_employees_excel_impl(brand='Vivo')

@app.route('/claro/export_employees_excel')
@login_required
def export_employees_excel_claro():
    return export_employees_excel_impl(brand='Claro')

def export_employees_excel_impl(brand='Vivo'):
    try:
        # Inicializar a query com filtro de marca
        query = Employee.query.filter_by(brand=brand)
        
        # Coletar e aplicar filtros de texto
        text_filters = ['registration', 'full_name', 'role', 'employee_type', 'status', 
                       'course_status', 'team', 'course_location', 'manager', 
                       'corporate_manager', 'instructor', 'operation_ready']
        
        for field in text_filters:
            value = request.args.get(field)
            if value and value != 'todos':  # 'todos' é o valor padrão dos selects
                # Se for um campo de seleção múltipla (pode conter valores separados por vírgula)
                if ',' in value:
                    values = [v.strip() for v in value.split(',') if v.strip()]
                    if values:
                        query = query.filter(getattr(Employee, field).in_(values))
                else:
                    query = query.filter(getattr(Employee, field).ilike(f'%{value}%'))
        
        # Aplicar filtros de data
        date_fields = [
            'admission_date', 'integration_start', 'integration_end', 
            'normative_start', 'normative_end', 'technical_course_start', 
            'technical_course_end', 'double_start', 'double_end', 
            'loading_date', 'field_operation_date'
        ]
        
        for date_field in date_fields:
            start_date = request.args.get(f'{date_field}_start')
            end_date = request.args.get(f'{date_field}_end')
            
            if start_date:
                try:
                    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                    query = query.filter(getattr(Employee, date_field) >= start_date)
                except (ValueError, TypeError):
                    pass
            
            if end_date:
                try:
                    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                    query = query.filter(getattr(Employee, date_field) <= end_date)
                except (ValueError, TypeError):
                    pass
        
        # Get filtered employees
        employees = query.all()
        
        # Convert to pandas DataFrame
        data = []
        for emp in employees:
            emp_data = {
                'Matrícula': emp.registration,
                'Nome Completo': emp.full_name,
                'Cargo': emp.role,
                'Tipo de Colaborador': emp.employee_type,
                'Data de Admissão': emp.admission_date.strftime('%d/%m/%Y') if emp.admission_date else '',
                'CEP': emp.cep,
                'Status': emp.status,
                'Status do Curso': emp.course_status,
                'Time': emp.team,
                'Local do Curso': emp.course_location,
                'Gestor': emp.manager,
                'Gestor Corporativo': emp.corporate_manager,
                'Instrutor': emp.instructor,
                'Contato': emp.contato,
                'Pronto para Operação': emp.operation_ready,
                'Início Integração': emp.integration_start.strftime('%d/%m/%Y') if emp.integration_start else '',
                'Fim Integração': emp.integration_end.strftime('%d/%m/%Y') if emp.integration_end else '',
                'Início Normativo': emp.normative_start.strftime('%d/%m/%Y') if emp.normative_start else '',
                'Fim Normativo': emp.normative_end.strftime('%d/%m/%Y') if emp.normative_end else '',
                'Início Curso Técnico': emp.technical_course_start.strftime('%d/%m/%Y') if emp.technical_course_start else '',
                'Fim Curso Técnico': emp.technical_course_end.strftime('%d/%m/%Y') if emp.technical_course_end else '',
                'Início Duplado': emp.double_start.strftime('%d/%m/%Y') if emp.double_start else '',
                'Fim Duplado': emp.double_end.strftime('%d/%m/%Y') if emp.double_end else '',
                'Data de Carregamento': emp.loading_date.strftime('%d/%m/%Y') if emp.loading_date else '',
                'Data de Operação no Campo': emp.field_operation_date.strftime('%d/%m/%Y') if emp.field_operation_date else '',
                'Fase Atual': emp.get_current_phase(),
                'Última Atualização': emp.last_updated.strftime('%d/%m/%Y %H:%M:%S') if emp.last_updated else ''
            }
            data.append(emp_data)
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Colaboradores')
            
            # Get the xlsxwriter workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Colaboradores']
            
            # Add a header format
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'top',
                'fg_color': '#D7E4BC',
                'border': 1
            })
            
            # Set column widths and format headers
            for i, col in enumerate(df.columns):
                # Set column width based on header length
                max_length = max(df[col].astype(str).apply(len).max(), len(col)) + 2
                worksheet.set_column(i, i, min(max_length, 30))
                
                # Write the column headers with the defined format
                worksheet.write(0, i, col, header_format)
        
        # Prepare the response
        output.seek(0)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return send_file(
            output,
            as_attachment=True,
            download_name=f'colaboradores_exportados_{timestamp}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return str(e), 500

@app.route('/delete_audit_logs', methods=['POST'])
@login_required
def delete_audit_logs():
    if current_user.access_type != 'admin':
        return jsonify({'success': False, 'error': 'Acesso negado'}), 403
    
    try:
        data = request.get_json()
        log_ids = data.get('ids', [])
        
        if not log_ids:
            return jsonify({'success': False, 'error': 'Nenhum ID de log fornecido'}), 400
        
        # Delete logs
        AuditLog.query.filter(AuditLog.id.in_(log_ids)).delete(synchronize_session=False)
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'{len(log_ids)} registros excluídos com sucesso'})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/audit_log')
@login_required
def audit_log():
    # Allow all authenticated users to view the audit log
    is_admin = current_user.access_type == 'admin'
    
    # Get filter parameters
    registration = request.args.get('registration')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    field = request.args.get('field')
    source = request.args.get('source')
    
    # Determinar a marca do usuário (se autenticado) para filtrar dados
    brand = getattr(current_user, 'brand', None) if current_user and hasattr(current_user, 'is_authenticated') and current_user.is_authenticated else None

    # Build query
    query = AuditLog.query
    # If brand is set, limit audit logs to registrations that belong to that brand
    if brand:
        regs_q = db.session.query(Employee.registration).filter(Employee.brand == brand).distinct()
        regs = [r[0] for r in regs_q.all()]
        if regs:
            query = query.filter(AuditLog.registration.in_(regs))
        else:
            # No registrations for this brand yet -> empty result
            query = query.filter(False)
    
    if registration:
        query = query.filter(AuditLog.registration == registration)
    if field and field != 'all':
        query = query.filter(AuditLog.field_changed == field)
    if source and source != 'all':
        query = query.filter(AuditLog.change_source == source)
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(AuditLog.changed_at >= start_date)
        except ValueError:
            pass
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(AuditLog.changed_at <= end_date)
        except ValueError:
            pass
    
    # Order by date descending
    logs = query.order_by(AuditLog.changed_at.desc()).all()
    
    # Get unique fields for filter dropdown
    fields = [r[0] for r in db.session.query(AuditLog.field_changed).distinct().all() if r[0]]
    
    return render_template('audit_log.html', 
                         logs=logs, 
                         fields=fields,
                         current_filters={
                             'registration': registration or '',
                             'start_date': start_date.strftime('%Y-%m-%d') if start_date and hasattr(start_date, 'strftime') else '',
                             'end_date': end_date.strftime('%Y-%m-%d') if end_date and hasattr(end_date, 'strftime') else '',
                             'field': field or 'all',
                             'source': source or 'all'
                         })


@app.route('/gestao_carregamento')
@login_required
def gestao_carregamento():
    # Obter apenas colaboradores aptos para operação e com data de carregamento definida
    # Determinar a marca do usuário (se autenticado) para filtrar dados
    brand = getattr(current_user, 'brand', None) if current_user and hasattr(current_user, 'is_authenticated') and current_user.is_authenticated else None

    # Obter apenas colaboradores aptos para operação e com data de carregamento definida
    colaboradores_q = Employee.query.filter(
        Employee.operation_ready == 'Sim',
        Employee.loading_date.isnot(None)
    )
    if brand:
        colaboradores_q = colaboradores_q.filter(Employee.brand == brand)
    colaboradores = colaboradores_q.order_by(Employee.loading_date).all()
    
    # Organizar por data de carregamento
    carregamentos = {}
    hoje = datetime.now().date()
    
    for colab in colaboradores:
        data_str = colab.loading_date.strftime('%Y-%m-%d')
        if data_str not in carregamentos:
            carregamentos[data_str] = {
                'data': colab.loading_date,
                'colaboradores': []
            }
        carregamentos[data_str]['colaboradores'].append(colab)
    
    # Separar em futuros e passados
    carregamentos_futuros = [v for k, v in carregamentos.items() if v['data'] >= hoje]
    carregamentos_passados = [v for k, v in carregamentos.items() if v['data'] < hoje]
    
    # Ordenar
    carregamentos_futuros.sort(key=lambda x: x['data'])
    carregamentos_passados.sort(key=lambda x: x['data'], reverse=True)
    
    return render_template(
        'gestao_carregamento.html',
        carregamentos_futuros=carregamentos_futuros,
        carregamentos_passados=carregamentos_passados,
        hoje=hoje
    )

@app.route('/exportar_carregamento')
@login_required
def exportar_carregamento():
    data = request.args.get('data')
    ids = request.args.get('ids')
    
    if not data and not ids:
        flash('Nenhum critério de exportação fornecido.', 'error')
        return redirect(url_for('gestao_carregamento'))
    
    # Construir a consulta base
    query = Employee.query.filter(Employee.operation_ready == 'Sim')
    
    # Filtrar por data ou IDs
    if data:
        try:
            data_filtro = datetime.strptime(data, '%Y-%m-%d').date()
            query = query.filter(Employee.loading_date == data_filtro)
            nome_arquivo = f'carregamento_{data}.xlsx'
        except ValueError:
            flash('Formato de data inválido.', 'error')
            return redirect(url_for('gestao_carregamento'))
    elif ids:
        try:
            ids_lista = [int(id) for id in ids.split(',')]
            query = query.filter(Employee.id.in_(ids_lista))
            nome_arquivo = f'carregamento_selecionados_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        except ValueError:
            flash('IDs inválidos fornecidos.', 'error')
            return redirect(url_for('gestao_carregamento'))
    
    # Executar a consulta
    colaboradores = query.order_by(Employee.full_name).all()
    
    if not colaboradores:
        flash('Nenhum colaborador encontrado com os critérios fornecidos.', 'warning')
        return redirect(url_for('gestao_carregamento'))
    
    # Criar DataFrame com os dados
    dados = []
    for colab in colaboradores:
        dados.append({
            'Matrícula': colab.registration,
            'Nome Completo': colab.full_name,
            'Função': colab.role or '',
            'Tipo': colab.employee_type or '',
            'Data Admissão': colab.admission_date.strftime('%d/%m/%Y') if colab.admission_date else '',
            'Turma': colab.team or '',
            'Data Carregamento': colab.loading_date.strftime('%d/%m/%Y') if colab.loading_date else '',
            'Local Frota': 'Informações por e-mail',
            'Local Almox': 'Informações por e-mail',
            'Gerente': colab.manager or '',
            'Gerente Corporativo': colab.corporate_manager or '',
            'Data Operação': colab.field_operation_date.strftime('%d/%m/%Y') if colab.field_operation_date else ''
        })
    
    # Criar Excel em memória
    output = BytesIO()
    df = pd.DataFrame(dados)
    df.to_excel(output, index=False, sheet_name='Carregamento')
    
    # Configurar resposta
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/apresentacao_duplado')
@login_required
def apresentacao_duplado():
    # Obter colaboradores com data de início de duplado definida
    # Mostrar apenas colaboradores aptos para operação (operation_ready ~ 'Sim')
    # Determinar a marca do usuário (se autenticado) para filtrar dados
    brand = getattr(current_user, 'brand', None) if current_user and hasattr(current_user, 'is_authenticated') and current_user.is_authenticated else None

    colaboradores_q = Employee.query.filter(
        Employee.double_start.isnot(None),
        db.or_(
            db.func.lower(Employee.operation_ready).like('sim%'),
            db.func.lower(Employee.operation_ready) == 's',
            db.func.lower(Employee.operation_ready) == '1',
            db.func.lower(Employee.operation_ready) == 'y',
            db.func.lower(Employee.operation_ready) == 'yes'
        )
    )
    if brand:
        colaboradores_q = colaboradores_q.filter(Employee.brand == brand)
    colaboradores = colaboradores_q.order_by(Employee.double_start).all()
    
    # Organizar por data de início do duplado
    duplados = {}
    hoje = datetime.now().date()
    
    for colab in colaboradores:
        data_str = colab.double_start.strftime('%Y-%m-%d')
        if data_str not in duplados:
            duplados[data_str] = {
                'data': colab.double_start,
                'colaboradores': []
            }
        duplados[data_str]['colaboradores'].append(colab)
    
    # Separar em futuros e passados
    duplados_futuros = [v for k, v in duplados.items() if v['data'] >= hoje]
    duplados_passados = [v for k, v in duplados.items() if v['data'] < hoje]
    
    # Ordenar
    duplados_futuros.sort(key=lambda x: x['data'])
    duplados_passados.sort(key=lambda x: x['data'], reverse=True)
    
    return render_template(
        'apresentacao_duplado.html',
        duplados_futuros=duplados_futuros,
        duplados_passados=duplados_passados,
        hoje=hoje
    )

@app.route('/exportar_duplado')
@login_required
def exportar_duplado():
    data = request.args.get('data')
    ids = request.args.get('ids')
    
    if not data and not ids:
        flash('Nenhum critério de exportação fornecido.', 'error')
        return redirect(url_for('apresentacao_duplado'))
    
    # Determinar a marca do usuário (se autenticado) para filtrar dados
    brand = getattr(current_user, 'brand', None) if current_user and hasattr(current_user, 'is_authenticated') and current_user.is_authenticated else None

    # Construir a consulta base (aplicar mesmo filtro de 'apto' usado na tela)
    query = Employee.query.filter(
        Employee.double_start.isnot(None),
        db.or_(
            db.func.lower(Employee.operation_ready).like('sim%'),
            db.func.lower(Employee.operation_ready) == 's',
            db.func.lower(Employee.operation_ready) == '1',
            db.func.lower(Employee.operation_ready) == 'y',
            db.func.lower(Employee.operation_ready) == 'yes'
        )
    )
    if brand:
        query = query.filter(Employee.brand == brand)
    
    # Filtrar por data ou IDs
    if data:
        try:
            data_filtro = datetime.strptime(data, '%Y-%m-%d').date()
            query = query.filter(Employee.double_start == data_filtro)
            nome_arquivo = f'duplado_{data}.xlsx'
        except ValueError:
            flash('Formato de data inválido.', 'error')
            return redirect(url_for('apresentacao_duplado'))
    elif ids:
        try:
            ids_lista = [int(id) for id in ids.split(',')]
            query = query.filter(Employee.id.in_(ids_lista))
            nome_arquivo = f'duplado_selecionados_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        except ValueError:
            flash('IDs inválidos fornecidos.', 'error')
            return redirect(url_for('apresentacao_duplado'))
    
    # Executar a consulta
    colaboradores = query.order_by(Employee.full_name).all()
    
    if not colaboradores:
        flash('Nenhum colaborador encontrado com os critérios fornecidos.', 'warning')
        return redirect(url_for('apresentacao_duplado'))
    
    # Criar DataFrame com os dados
    dados = []
    for colab in colaboradores:
        periodo_duplado = ""
        if colab.double_start:
            periodo_duplado = colab.double_start.strftime('%d/%m/%Y')
            if colab.double_end:
                periodo_duplado += f" a {colab.double_end.strftime('%d/%m/%Y')}"
        
        dados.append({
            'Matrícula': colab.registration,
            'Nome Completo': colab.full_name,
            'Função': colab.role or '',
            'Tipo': colab.employee_type or '',
            'Data Admissão': colab.admission_date.strftime('%d/%m/%Y') if colab.admission_date else '',
            'Turma': colab.team or '',
            'Período Duplado': periodo_duplado,
            'Local de Apresentação': 'Informações por e-mail',
            'Gerente': colab.manager or '',
            'Gerente Corporativo': colab.corporate_manager or '',
            'Data Operação': colab.field_operation_date.strftime('%d/%m/%Y') if colab.field_operation_date else ''
        })
    
    # Criar Excel em memória
    output = BytesIO()
    df = pd.DataFrame(dados)
    df.to_excel(output, index=False, sheet_name='Duplado')
    
    # Configurar resposta
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# Função para definir o bind dinamicamente

def set_bind_key(bind):
    for model in [Employee, User, AuditLog]:
        model.__bind_key__ = bind

@app.route('/')
def select_brand():
    return render_template('select_brand.html')

@app.route('/vivo/')
@login_required
def index_vivo():
    employees = Employee.query.filter_by(brand='Vivo').all()
    employees.sort(key=cmp_to_key(lambda a, b: safe_date_sort(a, b, 'field_operation_date')), reverse=True)
    return render_template('index.html', employees=employees, brand='Vivo')

@app.route('/claro/')
@login_required
def index_claro():
    employees = Employee.query.filter_by(brand='Claro').all()
    employees.sort(key=cmp_to_key(lambda a, b: safe_date_sort(a, b, 'field_operation_date')), reverse=True)
    return render_template('index.html', employees=employees, brand='Claro')

@app.route('/vivo/login', methods=['GET', 'POST'])
def login_vivo():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username, brand='Vivo').first()
        if user and check_password_hash(user.password, password):
            login_user(user)
            return redirect(url_for('index_vivo'))
        else:
            flash('Usuário ou senha inválidos')
    return render_template('login.html', brand='Vivo')

@app.route('/claro/login', methods=['GET', 'POST'])
def login_claro():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username, brand='Claro').first()
        if user and check_password_hash(user.password, password):
            login_user(user)
            return redirect(url_for('index_claro'))
        else:
            flash('Usuário ou senha inválidos')
    return render_template('login.html', brand='Claro')


# Custom unauthorized handler for Flask-Login
@login_manager.unauthorized_handler
def custom_unauthorized():
    path = request.path
    if path.startswith('/vivo/'):
        return redirect(url_for('login_vivo'))
    elif path.startswith('/claro/'):
        return redirect(url_for('login_claro'))
    else:
        # Se não for possível determinar a marca, redireciona para a seleção de marca
        return redirect(url_for('select_brand'))

@app.route('/ticket/<int:ticket_id>')
def ticket_override(ticket_id):
    """Temporary route to serve the fixed ticket template for testing visual fixes.

    Renders `ticket_fixed.html` from `templates/`. This is a testing helper
    so you can load `/ticket/1` locally and verify the CSS/JS adjustments.
    Remove this route after you're done testing.
    """
    try:
        return render_template('ticket_fixed.html')
    except Exception as e:
        return f"Error rendering ticket_fixed.html: {e}", 500


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        create_admin_user()
    # Use environment variable PORT to run instances on different ports when needed
    port = int(os.getenv('PORT', 5000))
    app.run(debug=True, host='0.0.0.0', port=port)
